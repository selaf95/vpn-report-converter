import streamlit as st
import pandas as pd
import re
from io import StringIO, BytesIO
from datetime import datetime
from fpdf import FPDF
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import os

# --- CLASE PDF ---
class CustomPDF(FPDF):
    def __init__(self, metadata):
        super().__init__()
        self.metadata = metadata

    def header(self):
        if os.path.exists("logo.jpg"):
            try:
                self.image("logo.jpg", x=10, y=8, w=30)
            except: pass
        self.set_y(20)
        self.set_font("helvetica", "B", 26)
        self.cell(0, 15, "System events", align="L", new_x="LMARGIN", new_y="NEXT")
        self.set_font("helvetica", "", 12)
        start, end = str(self.metadata.get('Start Date', '')), str(self.metadata.get('End Date', ''))
        self.cell(0, 10, f"{start} - {end}", align="L", new_x="LMARGIN", new_y="NEXT")
        self.ln(5)

    def footer(self):
        self.set_y(-25)
        self.set_font("helvetica", "I", 8)
        server_t = str(self.metadata.get('Server Time', ''))
        clean_f = f"Server time: {server_t}".encode('ascii', 'ignore').decode('ascii')
        self.cell(0, 10, clean_f, align='R')

# --- LÓGICA DE PROCESAMIENTO ---
def procesar_datos(uploaded_file):
    try:
        content = uploaded_file.getvalue().decode("utf-8")
    except:
        content = uploaded_file.getvalue().decode("latin-1")
    
    lines = content.splitlines()
    metadata = {}
    
    for line in lines[:15]:
        parts = [p.strip().replace('"', '') for p in line.split(',')]
        if len(parts) >= 2:
            key, val = parts[0], parts[1]
            if "Start Date" in key: metadata["Start Date"] = val
            elif "End Date" in key: metadata["End Date"] = val
            elif "Server Time" in key: metadata["Server Time"] = val
            elif "Appliance" in key and "Key" not in key: metadata["Appliance"] = val
            elif "Firmware Version" in key: metadata["Firmware Version"] = val
            elif "Device Serial Number" in key: metadata["Appliance Key"] = val
            elif "Criteria" in key or (len(parts) > 1 and "Event Type is" in parts[1]): 
                metadata["Criteria"] = val if "Criteria" not in key else parts[1]

    data_start = next((i for i, line in enumerate(lines) if "Time,Event Type,Severity,Message" in line), None)
    if data_start is None: return None, None, None

    df = pd.read_csv(StringIO('\n'.join(lines[data_start:])))
    df['Time'] = pd.to_datetime(df['Time'], errors='coerce')
    df = df[df['Message'].str.contains("SSL VPN User", na=False)]

    def extraer_usuario_accion(msg):
        match = re.search(r"SSL VPN User '([^']+)' (connected|disconnected)", msg)
        return match.groups() if match else (None, None)

    df[['Usuario', 'Accion']] = df['Message'].apply(extraer_usuario_accion).apply(pd.Series)
    df = df.dropna(subset=['Usuario', 'Accion']).sort_values(by=['Usuario', 'Time'])

    conexiones, activas = [], []
    for usuario, grupo in df.groupby('Usuario'):
        pila = []
        for _, fila in grupo.iterrows():
            if fila['Accion'] == 'connected':
                pila.append(fila['Time'])
            elif fila['Accion'] == 'disconnected' and pila:
                inicio = pila.pop(0)
                conexiones.append({
                    'Usuario': usuario, 'Inicio': inicio, 'Fin': fila['Time'], 
                    'Duración': str(fila['Time'] - inicio).split('.')[0]
                })
        for t in pila:
            activas.append({'Usuario': usuario, 'Inicio': t, 'Estado': 'Conectado'})

    return pd.DataFrame(conexiones), pd.DataFrame(activas), metadata

# --- INTERFAZ ---
archivo = st.file_uploader("Subir CSV de Sophos", type="csv")

if archivo:
    df_f, df_a, meta = procesar_datos(archivo)
    
    if meta:
        serial = meta.get("Appliance Key", "SERIAL")
        try:
            d_start = pd.to_datetime(meta.get("Start Date")).strftime("%Y-%m-%d")
            d_end = pd.to_datetime(meta.get("End Date")).strftime("%Y-%m-%d")
        except:
            d_start, d_end = "FECHA", "FECHA"

        nombre_base = f"Reporte_VPN_{serial}_{d_start}" if d_start == d_end else f"Reporte_VPN_{serial}_{d_start}_{d_end}"

        st.success(f"✅ Reporte generado")
        
        # --- EXCEL (NOMBRES DE PESTAÑA + CENTRADO + AUTOAJUSTE) ---
        output_excel = BytesIO()
        with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
            df_f.to_excel(writer, index=False, sheet_name='Conexiones completadas')
            df_a.to_excel(writer, index=False, sheet_name='Conexiones activas')
            
            for sheet_name in writer.sheets:
                ws = writer.sheets[sheet_name]
                for col in ws.columns:
                    max_length = 0
                    for cell in col:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    ws.column_dimensions[col[0].column_letter].width = max_length + 5

        # --- GENERACIÓN PDF ---
        try:
            pdf = CustomPDF(meta)
            pdf.add_page()
            pdf.set_font("helvetica", "", 10)
            pdf.cell(0, 7, f"Appliance: {meta.get('Appliance', 'N/A')}", new_x="LMARGIN", new_y="NEXT")
            pdf.cell(0, 7, f"Appliance key: {serial}", new_x="LMARGIN", new_y="NEXT")
            pdf.cell(0, 7, f"Firmware Version: {meta.get('Firmware Version', 'N/A')}", new_x="LMARGIN", new_y="NEXT")
            pdf.cell(0, 7, f"Filter(s) applied: {meta.get('Criteria', 'N/A')}", new_x="LMARGIN", new_y="NEXT")
            pdf.ln(5)

            # Tabla Conexiones Completadas
            pdf.set_font("helvetica", "B", 12)
            pdf.cell(0, 10, "Conexiones completadas", new_x="LMARGIN", new_y="NEXT")
            pdf.set_font("helvetica", "B", 9)
            col_widths = [60, 40, 40, 50]
            headers = ["Usuario", "Inicio", "Fin", "Duración"]
            for w, h in zip(col_widths, headers):
                pdf.cell(w, 8, h, border=1, align="C")
            pdf.ln()
            
            pdf.set_font("helvetica", "", 8)
            for _, r in df_f.iterrows():
                u = str(r["Usuario"]).encode('ascii', 'ignore').decode('ascii')
                # Dibujamos fila completa asegurando que no se corten las columnas
                pdf.cell(col_widths[0], 7, u, border=1)
                pdf.cell(col_widths[1], 7, r["Inicio"].strftime("%Y-%m-%d %H:%M"), border=1, align="C")
                pdf.cell(col_widths[2], 7, r["Fin"].strftime("%Y-%m-%d %H:%M"), border=1, align="C")
                pdf.cell(col_widths[3], 7, str(r["Duración"]), border=1, align="C")
                pdf.ln()

            # Tabla Conexiones Activas
            if not df_a.empty:
                pdf.ln(10)
                pdf.set_font("helvetica", "B", 12)
                pdf.cell(0, 10, "Conexiones activas", new_x="LMARGIN", new_y="NEXT")
                pdf.set_font("helvetica", "B", 9)
                pdf.cell(70, 8, "Usuario", border=1, align="C")
                pdf.cell(60, 8, "Inicio de Sesión", border=1, align="C")
                pdf.cell(60, 8, "Estado", border=1, align="C")
                pdf.ln()
                
                pdf.set_font("helvetica", "", 8)
                for _, r in df_a.iterrows():
                    u = str(r["Usuario"]).encode('ascii', 'ignore').decode('ascii')
                    pdf.cell(70, 7, u, border=1)
                    pdf.cell(60, 7, r["Inicio"].strftime("%Y-%m-%d %H:%M"), border=1, align="C")
                    pdf.cell(60, 7, "Conectado", border=1, align="C")
                    pdf.ln()

            # Fix para el error de descarga del PDF
            pdf_output = pdf.output()
            
            col1, col2 = st.columns(2)
            with col1:
                st.download_button("📥 Descargar Excel", output_excel.getvalue(), f"{nombre_base}.xlsx")
            with col2:
                st.download_button("📄 Descargar PDF", bytes(pdf_output), f"{nombre_base}.pdf")

        except Exception as e:
            st.error(f"Error en PDF: {e}")
            st.download_button("📥 Descargar Excel (Solo)", output_excel.getvalue(), f"{nombre_base}.xlsx")
